"""
Import the full EU CN8 nomenclature into app/data/cn_nomenclature.db.

Primary source format: Finnish Customs "CN YYYY official texts" XLSX
(Commission Implementing Regulation text, English descriptions).

Usage:
  python scripts/import_full_cn_nomenclature.py --download
  python scripts/import_full_cn_nomenclature.py --input app/data/sources/CN_2025_official_texts.xlsx
"""

from __future__ import annotations

import argparse
import re
import sqlite3
import sys
import time
import urllib.request
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.services.cn_database import (  # noqa: E402
    DEFAULT_DB_PATH,
    clear_data,
    cn_digits,
    init_schema,
    normalize_cn_code,
)

DEFAULT_DOWNLOAD_URL = (
    "https://tilastot.tulli.fi/documents/179508185/203434223/"
    "CN%202025%20official%20texts.xlsx/3fd063d3-48f0-b279-9d2d-63917e91658a?version=1.1"
)
DEFAULT_SOURCE_DIR = ROOT / "app" / "data" / "sources"

STOPWORDS = frozenset(
    {
        "a",
        "an",
        "and",
        "are",
        "as",
        "at",
        "be",
        "by",
        "for",
        "from",
        "in",
        "is",
        "it",
        "of",
        "on",
        "or",
        "the",
        "to",
        "with",
        "other",
        "than",
        "not",
        "elsewhere",
        "specified",
        "including",
        "having",
        "products",
        "product",
    }
)


def tokenize_keywords(*texts: str) -> str:
    tokens: set[str] = set()
    for text in texts:
        for token in re.findall(r"[a-z0-9]{3,}", text.lower()):
            if token not in STOPWORDS:
                tokens.add(token)
    return " ".join(sorted(tokens))


def digits_from_cn_cell(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\D", "", str(value).strip())


def description_from_dm(value: str | None) -> str:
    if not value:
        return ""
    text = str(value).strip()
    text = re.sub(r"^(SECTION|CHAPTER)\s+[IVXLC\d]+\s*-\s*", "", text, flags=re.I)
    text = re.sub(r"^CHAPTER\s+\d+\s*-\s*", "", text, flags=re.I)
    return text.strip()


def parse_finnish_official_texts(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = next((name for name in wb.sheetnames if name.upper().startswith("CN")), wb.sheetnames[0])
    ws = wb[sheet_name]

    ancestors: list[tuple[str, str]] = []
    records: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        _cnkey, cn_cell, _dashes, dm_cell = row[0], row[1], row[2], row[3]
        description = description_from_dm(dm_cell)
        if not description:
            continue

        cn_text = str(cn_cell).strip() if cn_cell is not None else ""
        digit_len = len(digits_from_cn_cell(cn_text))

        if digit_len == 0:
            level = "section" if "SECTION" in description.upper() else "chapter"
            ancestors.append((level, description))
            continue

        if digit_len <= 4:
            ancestors = [item for item in ancestors if item[0] in ("section", "chapter")]
            ancestors.append(("heading", description))
            continue

        if digit_len == 6:
            ancestors = [item for item in ancestors if item[0] != "subheading"]
            ancestors.append(("subheading", description))
            continue

        if digit_len < 8:
            continue

        digits = digits_from_cn_cell(cn_text)[:8]
        formatted = normalize_cn_code(digits)
        hierarchy_path = " > ".join(text for _level, text in ancestors[-6:] + [("cn8", description)])
        keywords = tokenize_keywords(hierarchy_path, description)
        records.append(
            {
                "cn_digits": digits,
                "cn_code": formatted,
                "description": description,
                "chapter_code": digits[:2],
                "heading_code": digits[:4],
                "hierarchy_path": hierarchy_path,
                "keywords": keywords,
            }
        )

    wb.close()
    return records


def download_source(url: str, dest: Path) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(url, headers={"User-Agent": "ExportGateway-CN-Import/1.0"})
    data = urllib.request.urlopen(request, timeout=120).read()
    dest.write_bytes(data)
    return dest


def import_records(
    records: list[dict],
    db_path: Path,
    *,
    source_label: str,
    source_url: str,
    cn_year: str,
) -> int:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()

    with sqlite3.connect(db_path) as conn:
        init_schema(conn)
        clear_data(conn)
        conn.executemany(
            """
            INSERT INTO cn_codes (
                cn_digits, cn_code, description, chapter_code, heading_code,
                hierarchy_path, keywords
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["cn_digits"],
                    row["cn_code"],
                    row["description"],
                    row["chapter_code"],
                    row["heading_code"],
                    row["hierarchy_path"],
                    row["keywords"],
                )
                for row in records
            ],
        )
        conn.executemany(
            """
            INSERT INTO cn_fts (cn_digits, cn_code, description, keywords, hierarchy_path)
            VALUES (?, ?, ?, ?, ?)
            """,
            [
                (
                    row["cn_digits"],
                    row["cn_code"],
                    row["description"],
                    row["keywords"],
                    row["hierarchy_path"],
                )
                for row in records
            ],
        )
        conn.execute(
            "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?)",
            ("source_label", source_label),
        )
        conn.execute(
            "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?)",
            ("source_url", source_url),
        )
        conn.execute(
            "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?)",
            ("cn_year", cn_year),
        )
        conn.execute(
            "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?)",
            ("record_count", str(len(records))),
        )
        conn.execute(
            "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?)",
            ("imported_at", time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())),
        )
        conn.commit()

    return len(records)


def main() -> int:
    parser = argparse.ArgumentParser(description="Import full EU CN8 nomenclature into SQLite.")
    parser.add_argument("--input", type=Path, help="Path to Finnish Customs CN official texts XLSX")
    parser.add_argument("--download", action="store_true", help="Download CN 2025 official texts from Finnish Customs")
    parser.add_argument("--url", default=DEFAULT_DOWNLOAD_URL, help="Override download URL")
    parser.add_argument("--output", type=Path, default=DEFAULT_DB_PATH, help="SQLite database output path")
    parser.add_argument("--year", default="2025", help="CN edition year label")
    args = parser.parse_args()

    if args.download:
        source_path = DEFAULT_SOURCE_DIR / f"CN_{args.year}_official_texts.xlsx"
        print(f"Downloading {args.url} ...")
        download_source(args.url, source_path)
        print(f"Saved {source_path.stat().st_size} bytes to {source_path}")
    elif args.input:
        source_path = args.input
    else:
        default = DEFAULT_SOURCE_DIR / f"CN_{args.year}_official_texts.xlsx"
        if default.is_file():
            source_path = default
        else:
            parser.error("Provide --input, place XLSX in app/data/sources/, or use --download")

    if not source_path.is_file():
        print(f"Source file not found: {source_path}", file=sys.stderr)
        return 1

    print(f"Parsing {source_path} ...")
    records = parse_finnish_official_texts(source_path)
    unique: dict[str, dict] = {}
    for row in records:
        unique[row["cn_digits"]] = row
    records = list(unique.values())
    records.sort(key=lambda item: item["cn_digits"])

    if len(records) < 5000:
        print(f"Warning: only {len(records)} CN8 codes parsed; expected ~10k+", file=sys.stderr)

    count = import_records(
        records,
        args.output,
        source_label="Finnish Customs — CN official texts (EU Commission regulation text, English)",
        source_url=args.url if args.download else str(source_path),
        cn_year=args.year,
    )
    print(f"Imported {count} CN8 codes into {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
